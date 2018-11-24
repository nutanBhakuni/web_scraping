import requests
from lxml import html
from time import sleep
import openpyxl

##def getbusinesslinks(base_url):
##    header = {'user-agent':'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36'}
##    urls = []
##    for page_number in range(1,13):
##        parameters = {
##                'search_terms':'movers and packers',
##                'geo_location_terms':'San Francisco, CA',
##                'page':page_number
##            }
##        response = requests.get(base_url,headers=header,params=parameters)
##        doc = html.fromstring(response.text)
##        print("Page Number ",page_number," : ")
##        for business in doc.xpath('//*[@id="main-content"]/div[@class="scrollable-pane"]/div[@class="search-results organic"]/div[@class="result"]'):
##            url = business.xpath('div/div[@class="v-card"]/div[@class="info"]/h2/a[@class="business-name"]/@href')[0]
##            urls.append(url)
##            print(url)
##    return urls
##    
##
##
##def writeToFile(businessLinks):
##    fp = open('businessLinks.txt','w')
##
##    for i in businessLinks:
##        fp.write(i+'\n')
##
##    fp.close()

def getfromfile():
    base_url = "https://www.yellowpages.com"
    urls = []
    fp = open('businessLinks.txt','r')
    for line in fp:
        urls.append(base_url+line)
    fp.close()
    return urls


def getbusinessdetails(businessLinks):
    header = {'user-agent':'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36'}
    temp_list = []
    for link in businessLinks:
        response = requests.get(link,headers=header)
        doc = html.fromstring(response.text)
        name = doc.xpath('//*[@id="main-header"]/article/div/h1/text()')[0].strip()

        try:
            address = doc.xpath('//*[@id="main-header"]/article/section[@class="primary-info"]/div[@class="contact"]/p[@class="address"]/span[1]/text()')[0].strip().rstrip(',')
        except:
            address = ""

        try:
            city = doc.xpath('//*[@id="main-header"]/article/section[@class="primary-info"]/div[@class="contact"]/p[@class="address"]/span[2]/text()')[0].strip().rstrip(',')
        except:
            city = ""

        try:
            state = doc.xpath('//*[@id="main-header"]/article/section[@class="primary-info"]/div[@class="contact"]/p[@class="address"]/span[3]/text()')[0].strip()
        except:
            state = ""

        try:
            zipcode = doc.xpath('//*[@id="main-header"]/article/section[@class="primary-info"]/div[@class="contact"]/p[@class="address"]/span[4]/text()')[0].strip()
        except:
            zipcode = ""

        try:
            phone = doc.xpath('//*[@id="main-header"]/article/section[@class="primary-info"]/div[@class="contact"]/p[@class="phone"]/text()')[0].strip()
        except:
            phone = ""
            
        print(name,"\t",address,"\t",city,"\t",state,"\t",zipcode,"\t",phone)
        temp = {
                "name":name,
                "address":address,
                "city":city,
                "state":state,
                "zipcode":zipcode,
                "phone":phone
                }
        temp_list.append(temp)
    return temp_list



def writeToFile(businessDetails):
    headings = ["Name","Address","City","State","Zipcode","Phone"]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Movers_And_Packers"
    row = 2
    
    for i in range(1,len(headings)+1):
        sheet.cell(1,i).value = headings[i-1]
    wb.save("D:\\nutan\\web scraping\\Yellow Pages\\Movers&PackersSanFranciscoCA.xlsx")

    for business in businessDetails:
        sheet.cell(row,1).value = business["name"]
        sheet.cell(row,2).value = business["address"]
        sheet.cell(row,3).value = business["city"]
        sheet.cell(row,4).value = business["state"]
        sheet.cell(row,5).value = business["zipcode"]
        sheet.cell(row,6).value = business["phone"]
        row = row + 1

    wb.save("D:\\nutan\\web scraping\\Yellow Pages\\Movers&PackersSanFranciscoCA.xlsx")


    
    
    
base_url = "https://www.yellowpages.com/search"
businessLinks = []
businessDetails = []
businessLinks = getfromfile()
businessDetails = getbusinessdetails(businessLinks)
writeToFile(businessDetails)
##businessLinks = getbusinesslinks(base_url)
##writeToFile(businessLinks)
